#!/usr/bin/env python3
"""Package the 13 star-schema CSVs into one multi-sheet .xlsx for Power BI
(Get Data -> Excel -> load the sheets you need). Data-only workbook (no
formulas), professional font, frozen+styled header rows, auto-fitted columns,
plus a Contents sheet documenting each table, the headline figures and the
methodology. Importable as build_workbook(); also runnable as a CLI."""
import argparse
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SAPPHIRE = "1B657C"; SIENNA = "EC6B4D"; ECRU = "F5F3E8"; BALTIC = "2C2C2C"

SHEETS = [
    ("fact_activities",       "fact_activities",       "One row per validated 4W entry, fully denormalised (gov + partner + indicator attributes, demographics, COD pcodes, correction flags)."),
    ("dim_sites",             "dim_sites",             "All masterlist sites + has_reported_activity flag (computed in pipeline)."),
    ("dim_partners",          "dim_partners",          "Partner directory: managing-in-masterlist and reporting-in-4W flags."),
    ("dim_indicators",        "dim_indicators",        "Canonical Activity Index (25 indicators) with counting class and framework category."),
    ("dim_geography",         "dim_geography",         "Governorate + neighborhood with integer COD pcodes (adm2 / adm4)."),
    ("dim_dates",             "dim_dates",             "Month dimension."),
    ("agg_partner_month",     "agg_partner_month",     "Partner x month: activities, unique sites, governorates, total reported."),
    ("agg_governorate_month", "agg_governorate_month", "Governorate x month: deduplicated reach, sites covered, masterlist denominators, coverage rate."),
    ("agg_indicator_month",   "agg_indicator_month",   "Indicator x month: activities, total, unique sites."),
    ("agg_neighborhood_month","agg_neighborhood_month","Neighborhood x month: deduplicated reach and sites covered."),
    ("agg_coverage",          "agg_coverage",          "Site x month coverage grid for gap analysis (covered yes/no)."),
    ("validation_log",        "validation_log",        "Row-level data-quality issues (QC ticket list to send back to partners)."),
    ("validation_summary",    "validation_summary",    "Issue counts per partner per month."),
]

MONTH_ABBR = {"January":"Jan","February":"Feb","March":"Mar","April":"Apr","May":"May",
              "June":"Jun","July":"Jul","August":"Aug","September":"Sep","October":"Oct",
              "November":"Nov","December":"Dec"}


def derive_period_label(out: Path) -> str:
    """e.g. 'Mar–Apr 2026' from dim_dates."""
    try:
        dd = pd.read_csv(out / "dim_dates.csv")
        months = list(dd.sort_values("reporting_period_id")["reporting_month"])
        year = int(dd["reporting_year"].iloc[0])
        abbr = [MONTH_ABBR.get(m, m[:3]) for m in months]
        if len(abbr) == 1:
            return f"{abbr[0]} {year}"
        return f"{abbr[0]}\u2013{abbr[-1]} {year}"
    except Exception:
        return ""


def build_workbook(outputs_dir, xlsx_path, period_label: str | None = None) -> dict:
    """Build the Power BI workbook from the CSVs in outputs_dir. Returns a small
    dict of headline figures. xlsx_path is where the workbook is written."""
    out = Path(outputs_dir)
    xlsx_path = str(xlsx_path)
    if period_label is None:
        period_label = derive_period_label(out)

    fact = pd.read_csv(out / "fact_activities.csv")
    dim_sites = pd.read_csv(out / "dim_sites.csv")
    rb = out / "_site_reach_readingB.csv"
    reach_b = int(pd.read_csv(rb)["reach"].sum()) if rb.exists() else None
    sites_reported = int(dim_sites["has_reported_activity"].sum())
    total_sites = len(dim_sites)
    n_valid = int(fact["is_valid"].sum())
    n_partners = int(fact.loc[fact["is_valid"], "organization"].nunique())

    # Promote the per-site Reading-B reach to a normal, documented table so the
    # de-duplicated headline reach is a simple column-sum in Power BI (no DAX).
    sheets = list(SHEETS)
    if rb.exists():
        site_reach = pd.read_csv(rb)
        # enrich with neighborhood + pcodes so this one table gives correct
        # cross-month reach at site, neighborhood OR governorate level
        ds = pd.read_csv(out / "dim_sites.csv")
        addc = [c for c in ["site_name", "neighborhood", "neighborhood_pcode",
                            "governorate_pcode"] if c in ds.columns]
        site_reach = site_reach.merge(ds[["site_id"] + addc], on="site_id", how="left")
        order = [c for c in ["site_id", "site_name", "governorate", "governorate_pcode",
                             "neighborhood", "neighborhood_pcode",
                             "capped", "cumulative", "reach"] if c in site_reach.columns]
        site_reach = site_reach[order]
        site_reach.to_csv(out / "agg_site_reach.csv", index=False)
        sheets.insert(6, ("agg_site_reach", "agg_site_reach",
            "Per-site de-duplicated reach (Reading B, cross-month). SUM of 'reach' = the "
            "cluster headline reach; group by governorate OR neighborhood for correct "
            "cross-month reach by area. Use THIS for any 'people reached' total — not the "
            "monthly tables, which double-count people served in more than one month."))

    present = [(s, stem, d) for s, stem, d in sheets if (out / f"{stem}.csv").exists()]

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as xw:
        pd.DataFrame({"_": [""]}).to_excel(xw, sheet_name="Contents", index=False)
        for sheet, stem, _d in present:
            df = pd.read_csv(out / f"{stem}.csv")
            for col in df.columns:
                if "pcode" in col.lower():
                    df[col] = pd.to_numeric(df[col], errors="coerce").astype("Int64")
            df.to_excel(xw, sheet_name=sheet[:31], index=False)

    wb = load_workbook(xlsx_path)
    hdr_fill = PatternFill("solid", fgColor=SAPPHIRE)
    hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    body_font = Font(name="Calibri", size=10)

    def style_sheet(ws):
        maxc = ws.max_column
        for c in range(1, maxc + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = hdr_fill; cell.font = hdr_font
            cell.alignment = Alignment(horizontal="left", vertical="center")
        for col in range(1, maxc + 1):
            letter = get_column_letter(col)
            w = len(str(ws.cell(row=1, column=col).value or ""))
            for r in range(2, min(ws.max_row, 200) + 1):
                v = ws.cell(row=r, column=col).value
                if v is not None:
                    w = max(w, len(str(v)))
                ws.cell(row=r, column=col).font = body_font
            ws.column_dimensions[letter].width = min(max(w + 2, 9), 52)
        ws.row_dimensions[1].height = 18
        ws.freeze_panes = "A2"

    for sheet, _s, _d in present:
        style_sheet(wb[sheet[:31]])

    cs = wb["Contents"]
    for row in list(cs.iter_rows()):
        for cell in row:
            cell.value = None
    cs.sheet_view.showGridLines = False
    title_font = Font(name="Calibri", bold=True, size=16, color=SAPPHIRE)
    sub_font = Font(name="Calibri", size=10, color=BALTIC)
    kpi_font = Font(name="Calibri", bold=True, size=11, color=SIENNA)
    h_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)

    plabel = f" ({period_label})" if period_label else ""
    cs["A1"] = f"Gaza Site Management Cluster — 4W Summary{plabel}"
    cs["A1"].font = title_font
    cs["A2"] = ("Star-schema data pack for Power BI. Get Data → Excel → select the sheets you need. "
                "Tables are pre-joined on site_id, reporting_period_id, activity_code, organization.")
    cs["A2"].font = sub_font; cs["A2"].alignment = Alignment(wrap_text=True)
    cs.merge_cells("A2:E2"); cs.row_dimensions[2].height = 30

    kpis = [
        (f"Cluster reach (Reading B, cumulative)", f"{reach_b:,}" if reach_b is not None else "n/a"),
        ("Sites with reported activity", f"{sites_reported:,} of {total_sites:,}"),
        ("Validated activity rows", f"{n_valid:,}"),
        ("Reporting partners (validated)", f"{n_partners}"),
    ]
    r = 4
    cs[f"A{r}"] = "Headline figures"; cs[f"A{r}"].font = Font(name="Calibri", bold=True, size=11, color=BALTIC); r += 1
    for label, val in kpis:
        cs[f"A{r}"] = label; cs[f"A{r}"].font = sub_font
        cs[f"C{r}"] = val; cs[f"C{r}"].font = kpi_font; r += 1

    r += 1
    cs[f"A{r}"] = "Sheet"; cs[f"B{r}"] = "Rows"; cs[f"C{r}"] = "Contents"
    for col in ("A", "B", "C"):
        cell = cs[f"{col}{r}"]; cell.fill = hdr_fill; cell.font = h_font
        cell.alignment = Alignment(horizontal="left")
    r += 1
    for sheet, stem, desc in present:
        nrows = len(pd.read_csv(out / f"{stem}.csv"))
        cs[f"A{r}"] = sheet; cs[f"A{r}"].font = Font(name="Calibri", size=10, bold=True, color=SAPPHIRE)
        cs[f"B{r}"] = nrows; cs[f"B{r}"].font = body_font
        cs[f"C{r}"] = desc; cs[f"C{r}"].font = body_font; cs[f"C{r}"].alignment = Alignment(wrap_text=True)
        cs.row_dimensions[r].height = 28; r += 1

    r += 1
    cs[f"A{r}"] = "Method note"; cs[f"A{r}"].font = Font(name="Calibri", bold=True, size=11, color=BALTIC); r += 1
    note = ("Cluster reach uses the agreed Reading-B method: per site x month, take the MAX across overlapping "
            "people-indicators (a person served by several activities is counted once); across months, take the MAX "
            "for population-capped indicators and SUM only for cumulative cohorts (Cash-for-Work, Training). People-capped "
            "indicators are capped at the masterlist site population; '# of sites' indicators are clamped to 1 per row. "
            "Partner attribution is by the Organization Name column, not the filename. Neighborhood pcodes come from the "
            "masterlist, with a governorate-scoped COD name lookup filling rows whose site could not be resolved. Reach is "
            "de-duplicated across partners, so the cluster total is lower than the sum of partner reaches.")
    cs[f"A{r}"] = note; cs[f"A{r}"].font = sub_font; cs[f"A{r}"].alignment = Alignment(wrap_text=True, vertical="top")
    cs.merge_cells(f"A{r}:E{r}"); cs.row_dimensions[r].height = 110

    cs.column_dimensions["A"].width = 26
    cs.column_dimensions["B"].width = 9
    cs.column_dimensions["C"].width = 78
    cs.column_dimensions["D"].width = 4
    cs.column_dimensions["E"].width = 4
    wb.move_sheet("Contents", -(len(wb.sheetnames) - 1))
    wb.active = 0
    wb.save(xlsx_path)
    return {"reach": reach_b, "sites_reported": sites_reported, "total_sites": total_sites,
            "valid_rows": n_valid, "partners": n_partners, "period": period_label,
            "sheets": wb.sheetnames}


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--outputs", type=Path, default=Path("outputs"))
    ap.add_argument("--xlsx", type=Path, default=None,
                    help="output workbook path; default SMC_4W_<period>_PowerBI.xlsx")
    ap.add_argument("--period", type=str, default=None)
    a = ap.parse_args()
    period = a.period if a.period is not None else derive_period_label(a.outputs)
    xlsx = a.xlsx or Path(f"SMC_4W_{period.replace(' ','').replace(chr(8211),'').replace('-','') or 'output'}_PowerBI.xlsx")
    info = build_workbook(a.outputs, xlsx, period)
    print("wrote", xlsx)
    print(info)
