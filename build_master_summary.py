#!/usr/bin/env python3
"""
master_summary builder
=======================
After the partner files are compiled, this produces a `master_summary`-style
workbook matching the colleague's format, but using THIS cluster's indicator
codes. It adds an `indicator_code_mapping` tab cross-walking the two code
systems (matched by indicator text).

Tabs produced (data-driven from the compiled outputs + the masterlist):
  indicator_code_mapping  this cluster's code  <->  colleague's code
  from_partners           compiled activity rows (colleague column layout)
  zite_managed            masterlist managing-agency x governorate + demographics
  zite_assessed           masterlist updating-agency x governorate + demographics
  partners_summary        submissions per partner per month
  activities_summary      per indicator: # reports, # partners
  ocha_pergovernorate     governorate demographic pivot
  ocha_summary            cluster overview (managed vs assessed)
  logframe                this cluster's indicator framework
  dim                     months present

Importable as build_master_summary(); also runnable as a CLI.

NOTE: zite_managed / zite_assessed reflect the CURRENT masterlist snapshot
(one file), not a monthly history; they are labelled with the compiled period.
"""
import argparse
import re
from difflib import SequenceMatcher
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SAPPHIRE = "1B657C"; SIENNA = "EC6B4D"; BALTIC = "2C2C2C"

GOV_PSE = {"North Gaza": "PSE255", "Gaza": "PSE260", "Deir Al-Balah": "PSE265",
           "Khan Younis": "PSE270", "Rafah": "PSE275"}
MONTH_NUM = {"January":1,"February":2,"March":3,"April":4,"May":5,"June":6,
             "July":7,"August":8,"September":9,"October":10,"November":11,"December":12}


def _norm(s):
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9 ]", " ", str(s).lower())).strip()


def _period_str(month, year):
    n = MONTH_NUM.get(str(month), 0)
    return f"{int(year)}-{n:02d}" if n else str(month)


# --------------------------------------------------------------- mapping -----
def build_indicator_mapping(indicators_path, colleague_codes_path):
    """Cross-walk this cluster's indicator codes to the colleague's, by text.
    Anchored on this cluster's indicators. Returns a tidy DataFrame."""
    ui = pd.read_csv(indicators_path)
    ui = ui[["Indicator_Code", "Indicators"]].dropna().drop_duplicates("Indicator_Code")
    coll = pd.read_csv(colleague_codes_path)
    coll["k"] = coll["indicator"].apply(_norm)
    rows = []
    for _, r in ui.iterrows():
        q = _norm(r["Indicators"])
        best, bcode, btext = 0.0, None, None
        for _, c in coll.iterrows():
            sc = SequenceMatcher(None, q, c["k"]).ratio()
            if sc > best:
                best, bcode, btext = sc, c["colleague_code"], c["indicator"]
        rows.append({
            "user_code": r["Indicator_Code"],
            "user_indicator": r["Indicators"],
            "colleague_code": bcode if best >= 0.6 else "",
            "colleague_indicator": btext if best >= 0.6 else "",
            "match_score": round(best, 2),
            "confidence": "high" if best >= 0.85 else ("medium" if best >= 0.6 else "REVIEW"),
        })
    return pd.DataFrame(rows).sort_values("user_code").reset_index(drop=True)


# --------------------------------------------------------- masterlist agg ----
def _zite_table(mst, agency_col, out_name):
    """Aggregate masterlist demographics by (governorate, agency)."""
    m = mst.copy()
    def s(*cols):
        present = [c for c in cols if c in m.columns]
        return m[present].apply(pd.to_numeric, errors="coerce").sum(axis=1, min_count=1) if present else 0
    m["_pop"] = pd.to_numeric(m.get("Total Inv"), errors="coerce")
    m["_boys"] = s("Males 0-5", "Males 6-17")
    m["_girls"] = s("Females 0-5", "Females 6-17")
    m["_men1859"] = pd.to_numeric(m.get("Males 18-60"), errors="coerce")
    m["_women1859"] = pd.to_numeric(m.get("Females 18-60"), errors="coerce")
    m["_em"] = pd.to_numeric(m.get("Males 60+"), errors="coerce")
    m["_ef"] = pd.to_numeric(m.get("Females 60+"), errors="coerce")
    m["_maleT"] = s("Males 0-5", "Males 6-17", "Males 18-60", "Males 60+")
    m["_femaleT"] = s("Females 0-5", "Females 6-17", "Females 18-60", "Females 60+")
    m["_men"] = s("Males 18-60", "Males 60+")
    m["_women"] = s("Females 18-60", "Females 60+")
    m["_pwd"] = pd.to_numeric(m.get("Persons with disabilities"), errors="coerce")
    m["_gov"] = m["Governorate"]
    g = (m.dropna(subset=[agency_col])
           .groupby(["_gov", agency_col], as_index=False)
           .agg(sites=("Site ID", "nunique"), total_population=("_pop", "sum"),
                male_total=("_maleT", "sum"), female_total=("_femaleT", "sum"),
                boys_u18=("_boys", "sum"), girls_u18=("_girls", "sum"),
                men_18_59=("_men1859", "sum"), women_18_59=("_women1859", "sum"),
                elderly_male_60p=("_em", "sum"), elderly_female_60p=("_ef", "sum"),
                pwd_total=("_pwd", "sum"), men=("_men", "sum"), women=("_women", "sum")))
    g = g.rename(columns={"_gov": "governorate", agency_col: out_name})
    return g


# ------------------------------------------------------------- builder -------
def build_master_summary(outputs_dir, masterlist_path, indicators_path,
                         colleague_codes_path, xlsx_path, period_label=""):
    out = Path(outputs_dir)
    fact = pd.read_csv(out / "fact_activities.csv")
    valid = fact[fact["is_valid"] == True].copy() if "is_valid" in fact.columns else fact.copy()
    mst = pd.read_csv(masterlist_path, low_memory=False)
    mapping = build_indicator_mapping(indicators_path, colleague_codes_path)
    user2coll = dict(zip(mapping["user_code"], mapping["colleague_code"]))

    # ---- from_partners (colleague layout, THIS cluster's codes) ----
    fp = pd.DataFrame({
        "source_file": fact.get("source_file", ""),
        "month": [_period_str(m, y) for m, y in zip(fact.get("reporting_month", ""),
                                                    fact.get("reporting_year", 0))],
        "organization": fact.get("organization", ""),
        "donor": fact.get("implementing_via", ""),
        "governorate": fact.get("governorate", ""),
        "governorate_pcode": fact.get("governorate", "").map(GOV_PSE).fillna(""),
        "neighborhood": fact.get("neighborhood", ""),
        "neighborhood_pcode": fact.get("neighborhood_pcode", ""),
        "site_id": fact.get("site_id", ""),
        "site_name_raw": fact.get("site_name_raw", ""),
        "activity_status": fact.get("activity_status", ""),
        "primary_activity": fact.get("primary_activity", ""),
        "sub_activity": fact.get("sub_activity", ""),
        "indicator": fact.get("indicator", ""),
        "indicator_code": fact.get("indicator_code", ""),       # THIS cluster's code
        "unit": fact.get("unit", ""),
        "total_count": fact.get("total_count", ""),
        "male": fact.get("male", ""), "female": fact.get("female", ""),
        "boys_under18": fact.get("boys_under18", ""), "girls_under18": fact.get("girls_under18", ""),
        "men_18_59": fact.get("men_18_59", ""), "women_18_59": fact.get("women_18_59", ""),
        "elderly_male_60plus": fact.get("elderly_male_60plus", ""),
        "elderly_female_60plus": fact.get("elderly_female_60plus", ""),
        "male_pwd": fact.get("male_pwd", ""), "female_pwd": fact.get("female_pwd", ""),
        "pwd_total": fact.get("pwd_total", ""),
        "activity_details": fact.get("activity_details", ""),
        "is_valid": fact.get("is_valid", ""),
    })
    # adult men/women (18+), matching the colleague's derived columns
    fp["men"] = (pd.to_numeric(fp["men_18_59"], errors="coerce").fillna(0)
                 + pd.to_numeric(fp["elderly_male_60plus"], errors="coerce").fillna(0)).astype(int)
    fp["women"] = (pd.to_numeric(fp["women_18_59"], errors="coerce").fillna(0)
                   + pd.to_numeric(fp["elderly_female_60plus"], errors="coerce").fillna(0)).astype(int)

    # ---- zite_managed / zite_assessed (current masterlist snapshot) ----
    zm = _zite_table(mst, "Managing Agency", "managing_agency")
    za_col = "Updating Agency" if "Updating Agency" in mst.columns else "Managing Agency"
    za = _zite_table(mst, za_col, "assessing_org")
    plabel = period_label or "current snapshot"
    zm.insert(0, "month", plabel); za.insert(0, "month", plabel)

    # ---- partners_summary (submissions per partner per month) ----
    psum = (valid.assign(month=[_period_str(m, y) for m, y in
                               zip(valid["reporting_month"], valid["reporting_year"])])
                 .pivot_table(index="organization", columns="month",
                              values="fact_id", aggfunc="count", fill_value=0))
    psum["Grand Total"] = psum.sum(axis=1)
    psum = psum.reset_index().rename(columns={"organization": "Partner"})

    # ---- activities_summary (per THIS cluster's indicator) ----
    asum = (valid.groupby(["indicator_code", "indicator"], as_index=False)
                 .agg(**{"# REPORTS": ("fact_id", "count"),
                         "# PARTNERS": ("organization", pd.Series.nunique)})
                 .rename(columns={"indicator_code": "CODE", "indicator": "INDICATOR"})
                 .sort_values("CODE"))

    # ---- ocha_pergovernorate + ocha_summary ----
    perc = (zm.groupby("governorate", as_index=False)
              .agg(total_population=("total_population", "sum"),
                   men=("men", "sum"), women=("women", "sum"),
                   boys_u18=("boys_u18", "sum"), girls_u18=("girls_u18", "sum"),
                   pwd_total=("pwd_total", "sum")))
    ocha_sum = pd.DataFrame({
        "scope": ["zite_managed", "zite_assessed"],
        "sites": [int(zm["sites"].sum()), int(za["sites"].sum())],
        "total_population": [zm["total_population"].sum(), za["total_population"].sum()],
        "men": [zm["men"].sum(), za["men"].sum()],
        "women": [zm["women"].sum(), za["women"].sum()],
        "boys_u18": [zm["boys_u18"].sum(), za["boys_u18"].sum()],
        "girls_u18": [zm["girls_u18"].sum(), za["girls_u18"].sum()],
        "pwd_total": [zm["pwd_total"].sum(), za["pwd_total"].sum()],
    })

    # ---- logframe (this cluster's framework) ----
    ui = pd.read_csv(indicators_path)
    logframe = ui[[c for c in ["Activity_Code", "Primary_Activity", "Sub_Activity",
                               "Indicators", "Indicator_Code", "Unit", "Indicator Purpose"]
                   if c in ui.columns]].copy()

    # ---- dim (months present) ----
    dim = pd.DataFrame({"month": sorted(set(fp["month"]))})

    sheets = {
        "indicator_code_mapping": mapping,
        "from_partners": fp,
        "zite_managed": zm,
        "zite_assessed": za,
        "partners_summary": psum,
        "activities_summary": asum,
        "ocha_pergovernorate": perc,
        "ocha_summary": ocha_sum,
        "logframe": logframe,
        "dim": dim,
    }

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as xw:
        for name, df in sheets.items():
            df.to_excel(xw, sheet_name=name[:31], index=False)

    # styling
    wb = load_workbook(xlsx_path)
    hdr_fill = PatternFill("solid", fgColor=SAPPHIRE)
    hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    body_font = Font(name="Calibri", size=10)
    review_fill = PatternFill("solid", fgColor="FCE8E2")
    for name in sheets:
        ws = wb[name[:31]]
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = hdr_fill; cell.font = hdr_font
            cell.alignment = Alignment(horizontal="left", vertical="center")
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            w = len(str(ws.cell(row=1, column=col).value or ""))
            for r in range(2, min(ws.max_row, 200) + 1):
                v = ws.cell(row=r, column=col).value
                if v is not None:
                    w = max(w, len(str(v)))
                ws.cell(row=r, column=col).font = body_font
            ws.column_dimensions[letter].width = min(max(w + 2, 9), 48)
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 18
    # highlight REVIEW rows in the mapping tab
    mw = wb["indicator_code_mapping"]
    conf_col = [c.column for c in mw[1] if c.value == "confidence"]
    if conf_col:
        cc = conf_col[0]
        for r in range(2, mw.max_row + 1):
            if mw.cell(row=r, column=cc).value == "REVIEW":
                for c in range(1, mw.max_column + 1):
                    mw.cell(row=r, column=c).fill = review_fill
    wb.save(xlsx_path)
    return {"tabs": list(sheets), "mapping_rows": len(mapping),
            "review": int((mapping["confidence"] == "REVIEW").sum()),
            "from_partners_rows": len(fp)}


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--outputs", type=Path, default=Path("outputs"))
    ap.add_argument("--masterlist", type=Path, required=True)
    ap.add_argument("--indicators", type=Path, required=True)
    ap.add_argument("--colleague-codes", type=Path, required=True)
    ap.add_argument("--xlsx", type=Path, default=Path("master_summary_output.xlsx"))
    ap.add_argument("--period", type=str, default="")
    a = ap.parse_args()
    info = build_master_summary(a.outputs, a.masterlist, a.indicators,
                                a.colleague_codes, a.xlsx, a.period)
    print("wrote", a.xlsx)
    print(info)
